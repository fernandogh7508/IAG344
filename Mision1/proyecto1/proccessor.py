import re
from openpyxl import load_workbook
# ================================
# FUNCION clean_id
# Elimina caracteres no númericos de un documento
# "cc75.888.56" = "7588856"
# ================================
def clean_id(value):
    if value is None:
        return ""
    return re.sub(r'\D','',str(value))
# ================================
#  FUNCION merge_name
#  Une nombre y aperllido en un solo campo
# ================================
def merge_name(name, lastname):
    if name is None:
        name =""
    if lastname is None:
        lastname =""
    return f"{name} {lastname}".strip()


def process_excel(path):
    #Acceso a la hoja llamada "datos"
    wb= load_workbook(path)
    ws = wb["Datos"]
    #Recorrer todas las filas desde la fila 2
    for row in range(2,ws.max_row+1):
        #Columna D: identificador limpio
        ws[f"D{row}"] =clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value
        )